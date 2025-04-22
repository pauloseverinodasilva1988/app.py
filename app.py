from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)

# Caminhos das planilhas
caminho_planilha_perdas = os.path.abspath('dados_perdas.xlsx')
caminho_planilha_produzidos = os.path.abspath('dados_produzidos.xlsx')
caminho_planilha_planejado = os.path.abspath('dados_planejado.xlsx')
caminho_planilha_unificada = os.path.abspath('dados_unificado.xlsx')

# Página inicial
@app.route('/inicio')
def inicio():
    return render_template('index.html')


@app.route('/', methods=['GET', 'POST'])
def formulario():
    campos = ["data", "codigo", "responsavel", "turno", "horario", "linha", "parte_processo", "quantidade", "tipo_de_perda"]

    if request.method == 'POST':
        dados = request.form.to_dict()

        # SALVAR NA PLANILHA DE PERDAS
        linha_perdas = [dados.get(campo, "").strip() for campo in campos]

        if not os.path.exists(caminho_planilha_perdas):
            wb = Workbook()
            ws = wb.active
            ws.append(campos)
        else:
            wb = load_workbook(caminho_planilha_perdas)
            ws = wb.active

        ws.append(linha_perdas)
        wb.save(caminho_planilha_perdas)

        # SALVAR NA PLANILHA UNIFICADA
        try:
            data_original = dados.get("data", "")
            data_formatada = datetime.strptime(data_original, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            data_formatada = data_original

        try:
            perda = float(dados.get("quantidade", 0))
        except ValueError:
            perda = 0

        parte_processo = dados.get("parte_processo", "")

        # Criando a linha para a planilha unificada, agora com "parte_processo"
        linha_unificada = [
            data_formatada,
            dados.get("linha", ""),
            dados.get("codigo", ""),
            dados.get("turno", ""),
            "perda",
            dados.get("tipo_de_perda", ""),
            "", "", "", "", "",
            perda,         # Coluna 'perda'
            parte_processo   # Coluna 'parte_processo'
        ]

        if not os.path.exists(caminho_planilha_unificada):
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.append([  # Adicionando "parte_processo" ao cabeçalho
                "data", "linha", "produto", "turno", "setor", "tipo_de_perda",
                "Peso", "qtd cx", "qtd kg", "prog", "prod", "perda", "perda_processo"
            ])
        else:
            wb2 = load_workbook(caminho_planilha_unificada)
            ws2 = wb2.active

        ws2.append(linha_unificada)
        wb2.save(caminho_planilha_unificada)

        return redirect(url_for('formulario', sucesso=1))

    sucesso = request.args.get('sucesso')
    return render_template('form.html', sucesso=sucesso)


# Formulário Produzidos
@app.route('/produzidos', methods=['GET', 'POST'])
def formulario_produzidos():
    if request.method == 'POST':
        dados = request.form.to_dict()

        fabricacao_br = ""
        if dados.get("fabricacao"):
            try:
                fabricacao = datetime.strptime(dados["fabricacao"], "%Y-%m-%d")
                fabricacao_br = fabricacao.strftime("%d/%m/%Y")
            except ValueError:
                fabricacao_br = dados["fabricacao"]

        try:
            peso_str = str(dados.get("peso", "0")).replace(",", ".").strip()
            peso = float(peso_str)
        except (ValueError, TypeError):
            peso = 0

        try:
            quantidade = int(dados.get("quantidade", 0))
        except (ValueError, TypeError):
            quantidade = 0

        qtd_kg = peso * quantidade

        dados_produzidos = [
            dados.get("codigo", "").strip(), peso, quantidade, fabricacao_br,
            dados.get("turno", "").strip(), dados.get("linha", "").strip()
        ]

        if not os.path.exists(caminho_planilha_produzidos):
            wb_prod = Workbook()
            ws_prod = wb_prod.active
            ws_prod.append(["codigo", "peso", "quantidade", "fabricacao", "turno", "linha"])
        else:
            wb_prod = load_workbook(caminho_planilha_produzidos)
            ws_prod = wb_prod.active

        ws_prod.append(dados_produzidos)
        wb_prod.save(caminho_planilha_produzidos)

        setor = "produzido"
        prog = ""
        prod = qtd_kg
        perda = ""

        unificado = [
            fabricacao_br, dados.get("linha", "").strip(), dados.get("codigo", "").strip(),
            dados.get("turno", "").strip(), setor, "-", peso, quantidade, qtd_kg, prog, prod, perda
        ]

        if not os.path.exists(caminho_planilha_unificada):
            wb_uni = Workbook()
            ws_uni = wb_uni.active
            ws_uni.append(["data", "linha", "produto", "turno", "setor", "problema",
                           "Peso", "qtd cx", "qtd kg", "prog", "prod", "perda"])
        else:
            wb_uni = load_workbook(caminho_planilha_unificada)
            ws_uni = wb_uni.active

        ws_uni.append(unificado)
        wb_uni.save(caminho_planilha_unificada)

        return redirect(url_for('formulario_produzidos', sucesso=1))

    sucesso = request.args.get('sucesso')
    return render_template('produzidos.html', sucesso=sucesso)


# Formulário Planejado (campo senha removido)
@app.route('/planejado', methods=['GET', 'POST'])
def formulario_planejado():
    if request.method == 'POST':
        # Processar os dados do formulário sem verificação de senha
        dados = request.form.to_dict()

        codigo = dados.get("codigo", "")
        fabricacao = dados.get("fabricacao", "")
        turno = dados.get("turno", "")
        linha = dados.get("linha", "")
        peso = float(dados.get("peso", 0))
        quantidade = int(dados.get("quantidade", 0))

        try:
            fabricacao_br = datetime.strptime(fabricacao, '%Y-%m-%d').strftime('%d/%m/%Y')
        except ValueError:
            fabricacao_br = fabricacao

        if not os.path.exists(caminho_planilha_planejado):
            wb = Workbook()
            ws = wb.active
            ws.append(["codigo", "fabricacao", "turno", "linha", "peso", "quantidade"])
        else:
            wb = load_workbook(caminho_planilha_planejado)
            ws = wb.active

        ws.append([codigo, fabricacao_br, turno, linha, peso, quantidade])
        wb.save(caminho_planilha_planejado)

        qtd_kg = peso * quantidade
        prog = qtd_kg

        unificado = [
            fabricacao_br, linha, codigo, turno, "programado", "-",
            peso, quantidade, qtd_kg, prog, "", ""
        ]

        if not os.path.exists(caminho_planilha_unificada):
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.append(["data", "linha", "produto", "turno", "setor", "problema",
                        "Peso", "qtd cx", "qtd kg", "prog", "prod", "perda"])
        else:
            wb2 = load_workbook(caminho_planilha_unificada)
            ws2 = wb2.active

        ws2.append(unificado)
        wb2.save(caminho_planilha_unificada)

        return redirect(url_for('formulario_planejado', sucesso=1))

    sucesso = request.args.get('sucesso')
    return render_template('planejado.html', sucesso=sucesso)


# Tabela unificada
@app.route('/tabela-unificada')
def tabela_unificada():
    registros = []
    if os.path.exists(caminho_planilha_unificada):
        wb = load_workbook(caminho_planilha_unificada)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            registros.append(row)
    return render_template('tabela_unificada.html', registros=registros)

@app.route('/tabela')
def tabela_perdas():
    registros = []
    if os.path.exists(caminho_planilha_perdas):
        wb = load_workbook(caminho_planilha_perdas)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            registros.append(row)
    return render_template('tabela.html', registros=registros, titulo='Tabela de Perdas')

@app.route('/tabela-produzidos')
def tabela_produzidos():
    registros = []
    if os.path.exists(caminho_planilha_produzidos):
        wb = load_workbook(caminho_planilha_produzidos)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            registros.append(row)
    return render_template('tabela_produzidos.html', registros=registros)

@app.route('/tabela-planejado')
def tabela_planejado():
    registros = []
    if os.path.exists(caminho_planilha_planejado):
        wb = load_workbook(caminho_planilha_planejado)
        ws = wb.active
        cabecalho_planejado = [cell.value for cell in ws[1]]
        registros.append(cabecalho_planejado)  # Adiciona o cabeçalho
        for row in ws.iter_rows(min_row=2, values_only=True):
            registros.append(row)
    return render_template('tabela_planejado.html', registros=registros)

@app.route('/dash')
def dashboard():
    total_prog = 0
    total_prod = 0
    total_perda = 0
    perda_por_tipo = {}
    total_peso_produzido = 0
    total_peso_planejado = 0
    perdas_detalhadas = []

    # Carregar parte_processo da planilha de perdas
    mapa_linha_para_parte = {}
    if os.path.exists(caminho_planilha_perdas):
        wb_perdas = load_workbook(caminho_planilha_perdas)
        ws_perdas = wb_perdas.active
        cabecalho_perdas = [cell.value for cell in ws_perdas[1]]
        try:
            idx_linha = cabecalho_perdas.index('linha')
            idx_parte_processo = cabecalho_perdas.index('parte_processo')

            for row in ws_perdas.iter_rows(min_row=2, values_only=True):
                linha = row[idx_linha]
                parte = row[idx_parte_processo]
                if linha:
                    mapa_linha_para_parte[linha] = parte
        except ValueError as e:
            print(f"Erro ao encontrar colunas no cabeçalho de perdas: {e}")

    # Carregar dados da planilha unificada
    if os.path.exists(caminho_planilha_unificada):
        wb = load_workbook(caminho_planilha_unificada)
        ws = wb.active
        cabecalho = [cell.value for cell in ws[1]]
        try:
            col_linha = cabecalho.index('linha')
            col_setor = cabecalho.index('setor')
            col_tipo_perda = cabecalho.index('tipo_de_perda')
            col_prog = cabecalho.index('prog')
            col_prod = cabecalho.index('prod')
            col_perda = cabecalho.index('perda')
            col_peso = cabecalho.index('Peso')
            col_qtd_cx = cabecalho.index('qtd cx') if 'qtd cx' in cabecalho else -1
            col_quantidade = cabecalho.index('quantidade') if 'quantidade' in cabecalho else -1
        except ValueError as e:
            print(f"Erro ao encontrar colunas no cabeçalho unificado: {e}")
            return render_template('dash.html', total_prog=0, total_prod=0, eficiencia=0, total_perda=0,
                                   perda_pct=0, perda_por_setor=[], rendimento=0, atendimento_plano=0,
                                   perdas_detalhadas=[])

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                linha = row[col_linha]
                setor = row[col_setor]
                tipo_perda = row[col_tipo_perda]
                prog_val = float(row[col_prog]) if row[col_prog] is not None else 0
                prod_val = float(row[col_prod]) if row[col_prod] is not None else 0
                perda_val = float(row[col_perda]) if row[col_perda] is not None else 0
                peso_val = float(row[col_peso]) if row[col_peso] is not None else 0
                qtd_cx_val = float(row[col_qtd_cx]) if col_qtd_cx != -1 and row[col_qtd_cx] is not None else 0
                quantidade_val = float(row[col_quantidade]) if col_quantidade != -1 and row[col_quantidade] is not None else 0

                if setor == 'programado':
                    peso_planejado_linha = peso_val * (qtd_cx_val if qtd_cx_val > 0 else quantidade_val)
                    total_peso_planejado += peso_planejado_linha
                    total_prog += prog_val
                elif setor == 'produzido':
                    peso_produzido_linha = peso_val * (qtd_cx_val if qtd_cx_val > 0 else quantidade_val)
                    total_peso_produzido += peso_produzido_linha
                    total_prod += prod_val
                elif setor == 'perda' and perda_val > 0 and tipo_perda:
                    perda_por_tipo[tipo_perda] = perda_por_tipo.get(tipo_perda, 0) + perda_val
                    total_perda += perda_val

                    perdas_detalhadas.append({
                        'linha': linha,
                        'parte_processo': mapa_linha_para_parte.get(linha, 'Desconhecido'),
                        'tipo_perda': tipo_perda,
                        'quantidade': perda_val
                    })
            except (ValueError, IndexError) as e:
                print(f"Erro detalhado ao processar linha: {e}, Row Data: {row}")
                continue

    eficiencia = (total_prod / total_prog) * 100 if total_prog > 0 else 0
    perda_pct = (total_perda / (total_prod + total_perda)) * 100 if (total_prod + total_perda) > 0 else 0
    rendimento = (total_prod / (total_prod + total_perda)) * 100 if (total_prod + total_perda) > 0 else 0
    atendimento_plano = (total_prod / total_prog) * 100 if total_prog > 0 else 0

    perda_para_template = [{'tipo_de_perda': tipo, 'quantidade': quantidade} for tipo, quantidade in perda_por_tipo.items()]

    return render_template('dash.html',
                           total_prog=total_prog,
                           total_prod=total_prod,
                           eficiencia=eficiencia,
                           total_perda=total_perda,
                           perda_pct=perda_pct,
                           perda_por_setor=perda_para_template,
                           rendimento=rendimento,
                           atendimento_plano=atendimento_plano,
                           perdas_detalhadas=perdas_detalhadas)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
